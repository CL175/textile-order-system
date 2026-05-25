# -*- coding: utf-8 -*-
"""Convert .xls to .xlsx via xlrd + openpyxl (no COM dependency)."""

import os, sys
import xlrd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def convert_one(src_path):
    book = xlrd.open_workbook(src_path, formatting_info=True)
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    for sheet_name in book.sheet_names():
        old_ws = book.sheet_by_name(sheet_name)
        new_ws = wb.create_sheet(title=sheet_name)

        for r in range(old_ws.nrows):
            for c in range(old_ws.ncols):
                val = old_ws.cell_value(r, c)
                ctype = old_ws.cell_type(r, c)
                if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_ERROR):
                    continue
                cell = new_ws.cell(row=r + 1, column=c + 1)
                if ctype == xlrd.XL_CELL_DATE:
                    import datetime
                    dt_tuple = xlrd.xldate_as_tuple(val, book.datemode)
                    try:
                        cell.value = datetime.datetime(*dt_tuple)
                    except:
                        cell.value = val
                else:
                    cell.value = val

        # column widths
        for c in range(old_ws.ncols):
            ci = old_ws.colinfo_map.get(c)
            if ci and ci.width:
                w = ci.width
                if w > 100:
                    w = w / 256.0
                new_ws.column_dimensions[get_column_letter(c + 1)].width = w

        # row heights
        for r in range(old_ws.nrows):
            ri = old_ws.rowinfo_map.get(r)
            if ri and ri.height:
                new_ws.row_dimensions[r + 1].height = ri.height / 20.0

        # merged cells
        for rlo, rhi, clo, chi in old_ws.merged_cells:
            new_ws.merge_cells(start_row=rlo + 1, start_column=clo + 1,
                               end_row=rhi, end_column=chi)

    dst = src_path.replace(".xls", ".xlsx")
    if dst == src_path:
        dst = src_path + ".xlsx"
    wb.save(dst)
    wb.close()
    return dst


def main():
    if len(sys.argv) < 2:
        print("Usage: python convert_xls_to_xlsx.py <dir>")
        sys.exit(1)

    target = sys.argv[1]
    if not os.path.exists(target):
        print("NOT FOUND:", target)
        sys.exit(1)

    files = []
    if os.path.isfile(target):
        if target.endswith(".xls"):
            files.append(target)
    else:
        for root, dirs, fns in os.walk(target):
            for fn in fns:
                if fn.endswith(".xls") and ".bak" not in fn:
                    files.append(os.path.join(root, fn))

    print("Found {} .xls file(s)".format(len(files)))
    ok = 0
    for fp in files:
        print("  {} ...".format(os.path.basename(fp)), end=" ")
        try:
            dst = convert_one(fp)
            # backup original
            bak = fp + ".bak"
            n = 1
            while os.path.exists(bak):
                n += 1
                bak = fp + ".bak{}".format(n)
            os.rename(fp, bak)
            print("OK ->", os.path.basename(dst))
            ok += 1
        except Exception as e:
            print("FAIL:", e)
    print("\nDone: {} OK, {} failed".format(ok, len(files) - ok))


if __name__ == "__main__":
    main()
