# -*- coding: utf-8 -*-
"""Delivery number generation.

Format: {customer_prefix}{YY}{MM}{3-digit-seq}
Example: HR202605001

Sequence is per-customer per-month, independent of order creation.
Only assigned when generating a delivery note.

The generator also scans all sheets in the customer's Excel delivery file
(cell B4 of each sheet) to detect manually bumped-up delivery numbers
and continue from the highest one, so that Excel edits are not lost.
"""
import io
import os
import re

from datetime import datetime
from db.database import get_connection


def generate_delivery_number(customer_id, delivery_date=None):
    """Generate the next delivery number for a customer.

    Args:
        customer_id: customer ID
        delivery_date: date string "YYYY-MM-DD", defaults to today

    Returns:
        str: delivery number like "HR202605001", or None on error
    """
    if delivery_date is None:
        today = datetime.now()
    else:
        parts = delivery_date.split("-")
        today = datetime(int(parts[0]), int(parts[1]), int(parts[2]))

    yyyy = today.strftime("%Y")
    mm = today.strftime("%m")
    prefix = _get_customer_prefix(customer_id)
    if not prefix:
        return None
    like_pattern = "{}{}{}%".format(prefix, yyyy, mm)

    c = get_connection()
    try:
        row = c.execute(
            "SELECT delivery_number FROM delivery_note"
            " WHERE delivery_number LIKE ?"
            " ORDER BY delivery_number DESC LIMIT 1",
            (like_pattern,)).fetchone()

        if row:
            last_seq = int(row["delivery_number"][-3:])
            seq = last_seq + 1
        else:
            seq = 1

        # ---- Excel 反查：扫描所有 sheet 的 B4 取最大单号 ----
        excel_seq = _read_excel_latest_seq(customer_id, prefix, yyyy, mm)
        if excel_seq is not None and excel_seq >= seq:
            seq = excel_seq + 1

        return "{}{}{}{:03d}".format(prefix, yyyy, mm, seq)
    finally:
        c.close()


def _read_excel_latest_seq(customer_id, prefix, year, month):
    """Scan all sheets in the customer's Excel file for the current month's
    highest delivery-number sequence.
    """
    try:
        from db.models import Settings
        cust_name = _get_customer_name(customer_id)
        if not cust_name:
            return None

        folder = Settings.get("default_excel_folder") or "D:/xxm/送货单"
        path = os.path.join(os.path.normpath(folder),
                            u"{}.xlsx".format(cust_name))
        if not os.path.exists(path):
            return None

        from openpyxl import load_workbook

        # Read into memory via shared-read to bypass WPS/Excel file lock
        try:
            with open(path, 'rb') as f:
                file_content = io.BytesIO(f.read())
        except Exception as e:
            print(f"无法读取Excel文件，可能权限异常: {e}")
            return None

        wb = load_workbook(file_content, data_only=True)

        try:
            pat = re.escape(prefix) + r'(\d{4})(\d{2})(\d+)'
            best = 0
            for name in wb.sheetnames:
                try:
                    val = wb[name]["B4"].value
                except Exception:
                    continue
                if not val or not isinstance(val, str):
                    continue

                m = re.search(pat, val, flags=re.IGNORECASE)
                if not m:
                    continue
                ey, em, es = m.group(1), m.group(2), m.group(3)
                if ey == year and em == month:
                    best = max(best, int(es))
            return best if best > 0 else None
        finally:
            wb.close()
    except Exception as e:
        print(f"解析Excel最新单号时发生异常: {e}")
        return None


def _get_customer_name(customer_id):
    c = get_connection()
    try:
        row = c.execute(
            "SELECT name FROM customer WHERE id=?",
            (customer_id,)).fetchone()
        return row["name"] if row else None
    finally:
        c.close()


def _get_customer_prefix(customer_id):
    c = get_connection()
    try:
        row = c.execute(
            "SELECT prefix FROM customer WHERE id=?",
            (customer_id,)).fetchone()
        return row["prefix"] if row else None
    finally:
        c.close()
