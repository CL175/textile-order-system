# -*- coding: utf-8 -*-
"""
Export: fill delivery note data into customer's delivery Excel file.

Strategy:
  1. First export for a customer: fill the template -> save as the delivery file.
  2. Subsequent exports: use openpyxl's ``copy_worksheet`` to duplicate an
     existing sheet **within the same workbook**, then clear the data area
     and fill with new data.

This guarantees 100% formatting fidelity because ``copy_worksheet`` operates
on the same ``styles.xml`` — no cross-workbook style remapping, no ZIP
manipulation, no WPS automation.  All operations happen in memory.
"""
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from db.models import DeliveryNote, DNItem, Customer, Settings


# ---------------------------------------------------------------------------
# Public API – called from dn_window.py
# ---------------------------------------------------------------------------
def generate_delivery_note(dn_id):
    """Generate (or append) a delivery note Excel sheet for the given DN.

    Returns:
        ``(True, filepath)`` on success, ``(False, error_message)`` on failure.
    """
    dn = DeliveryNote.get_by_id(dn_id)
    if not dn:
        return False, u"送货单不存在"

    items = DNItem.get_by_dn(dn_id)
    customer = Customer.get_by_id(dn["customer_id"])
    if not customer:
        return False, u"客户不存在"

    # Locate template
    tpl_path = os.path.join(_template_dir(),
                            u"{}.xlsx".format(customer["name"]))
    if not os.path.exists(tpl_path):
        return False, u"模板文件不存在:\n{}".format(tpl_path)

    # Delivery file path
    folder = os.path.normpath(
        Settings.get("default_excel_folder") or "D:/xxm/送货单")
    if not os.path.isdir(folder):
        os.makedirs(folder)
    delivery_path = os.path.join(folder,
                                 u"{}.xlsx".format(customer["name"]))

    sheet_name = _sheet_name(dn.get("delivery_date", ""))

    try:
        # Read template layout (right-label positions) once per export.
        # This is fast (<1ms) and ensures B4/B5 spacing matches the template.
        layout = _read_template_layout(tpl_path)

        if not os.path.exists(delivery_path):
            _create_from_template(tpl_path, delivery_path, dn, items,
                                  customer, sheet_name, layout)
        else:
            _append_sheet(delivery_path, dn, items, customer, sheet_name,
                          layout)

        return True, delivery_path

    except Exception:
        import sys
        import traceback
        from logic.logger import log_error
        log_error(*sys.exc_info(), extra_info="export dn {}".format(dn_id))
        return False, u"导出失败:\n{}".format(traceback.format_exc())


# ---------------------------------------------------------------------------
# First export: template -> new delivery file
# ---------------------------------------------------------------------------
def _create_from_template(tpl_path, delivery_path, dn, items, customer,
                          sheet_name, layout):
    wb = load_workbook(tpl_path)
    try:
        ws = wb.worksheets[0]
        ws.title = sheet_name
        _clear_data_area(ws)
        _fill_data(ws, dn, items, customer, layout)
        wb.save(delivery_path)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Subsequent exports: copy sheet within workbook -> fill data
# ---------------------------------------------------------------------------
def _append_sheet(delivery_path, dn, items, customer, desired_name, layout):
    wb = load_workbook(delivery_path)
    try:
        # Use the first sheet as formatting source (it was originally a
        # filled template, so its formatting is exactly what we need).
        source_ws = wb.worksheets[0]
        target_ws = wb.copy_worksheet(source_ws)

        # Resolve name (avoid duplicates)
        actual_name = desired_name
        if actual_name in wb.sheetnames:
            n = 2
            while u"{}({})".format(desired_name, n) in wb.sheetnames:
                n += 1
            actual_name = u"{}({})".format(desired_name, n)
        target_ws.title = actual_name

        _clear_data_area(target_ws)
        _fill_data(target_ws, dn, items, customer, layout)

        # Copy page setup from source (print area, margins, orientation, etc.)
        _copy_page_setup(source_ws, target_ws)

        wb.save(delivery_path)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Template layout detection
# ---------------------------------------------------------------------------
def _read_template_layout(tpl_path):
    """Extract the right-label positions from the template's B4/B5 cells.

    Returns a dict with keys ``b4_right`` and ``b5_right`` — the character
    position where the right-side label begins in each merged header cell.
    """
    wb = load_workbook(tpl_path)
    try:
        ws = wb.worksheets[0]
        b4_val = ws["B4"].value or ""
        b5_val = ws["B5"].value or ""

        # Find where the right-side keyword starts in B4
        b4_right = 44  # fallback
        for kw in [u"送货单编号：", u"送货单编号:", u"送货单编号",
                   u"客户订单号：", u"客户订单号:", u"客户订单号"]:
            idx = b4_val.find(kw)
            if idx > 0:
                b4_right = idx
                break

        # Find where the right-side keyword starts in B5
        b5_right = 42  # fallback
        for kw in [u"送货日期  ：", u"送货日期：", u"送货日期:", u"送货日期"]:
            idx = b5_val.find(kw)
            if idx > 0:
                b5_right = idx
                break

        return {"b4_right": b4_right, "b5_right": b5_right}
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Sheet name helpers
# ---------------------------------------------------------------------------
def _sheet_name(date_str):
    """Convert '2026-05-19' -> '5-19'."""
    if not date_str:
        import datetime
        d = datetime.date.today()
        return "{}-{}".format(d.month, d.day)
    parts = date_str.split("-")
    if len(parts) >= 2:
        try:
            return "{}-{}".format(int(parts[1]), int(parts[2]))
        except ValueError:
            pass
    return date_str


def _copy_page_setup(source_ws, target_ws):
    """Copy page setup settings from source to target worksheet.

    openpyxl's copy_worksheet does NOT copy page setup,
    so we do it manually here.
    """
    try:
        sps = source_ws.page_setup
        tps = target_ws.page_setup
        # orientation
        if sps.orientation:
            tps.orientation = sps.orientation
        # paper size
        if sps.paperSize:
            tps.paperSize = sps.paperSize
        # fit to
        if sps.fitToHeight:
            tps.fitToHeight = sps.fitToHeight
        if sps.fitToWidth:
            tps.fitToWidth = sps.fitToWidth
        # print area
        if source_ws.print_area:
            target_ws.print_area = source_ws.print_area
        # page margins
        sm = source_ws.page_margins
        tm = target_ws.page_margins
        if sm.top is not None:
            tm.top = sm.top
        if sm.bottom is not None:
            tm.bottom = sm.bottom
        if sm.left is not None:
            tm.left = sm.left
        if sm.right is not None:
            tm.right = sm.right
        if sm.header is not None:
            tm.header = sm.header
        if sm.footer is not None:
            tm.footer = sm.footer
        # sheet print options
        spo = source_ws.sheet_properties.pageSetUpPr
        tpo = target_ws.sheet_properties.pageSetUpPr
        if spo:
            tpo.fitToPage = spo.fitToPage
        # header/footer
        if source_ws.oddHeader:
            target_ws.oddHeader = source_ws.oddHeader
        if source_ws.oddFooter:
            target_ws.oddFooter = source_ws.oddFooter
        # print title rows/cols
        if source_ws.print_title_rows:
            target_ws.print_title_rows = source_ws.print_title_rows
        if source_ws.print_title_cols:
            target_ws.print_title_cols = source_ws.print_title_cols
    except Exception:
        pass  # page setup copy is best-effort


def _template_dir():
    return Settings.get("template_dir", "D:/xxm/templates")


# ---------------------------------------------------------------------------
# Cell / data helpers
# ---------------------------------------------------------------------------
def _clear_data_area(ws):
    """Erase the data rows (7-18, columns B-I)."""
    for r in range(7, 19):
        for c in range(2, 10):
            ws["{}{}".format(get_column_letter(c), r)].value = None


def _fill_data(ws, dn, items, customer, layout):
    """Write delivery note header and line items into the worksheet."""
    # B4: company email ...................... 客户订单号：DN-XXX
    b4_left = _clean_str(Settings.get("company_email")) or ""
    b4_right_label = u"送货单编号："
    b4_right_val = _clean_str(dn.get("delivery_number", "")) or ""
    _set(ws, "B4", _format_header(b4_left, b4_right_label + b4_right_val,
                                  layout["b4_right"]), _clean=False)

    # B5: 客户：NAME ................................ 送货日期：YYYY-MM-DD
    b5_left = u"客户：" + (_clean_str(customer.get("name", "")) or "")
    b5_right_label = u"送货日期："
    b5_right_val = _clean_str(dn.get("delivery_date", "")) or ""
    _set(ws, "B5", _format_header(b5_left, b5_right_label + b5_right_val,
                                  layout["b5_right"]), _clean=False)

    # Determine column layout from customer's dn_headers.
    # Column order: 客户号, 订单号, 制单号, 款号
    # Data mapping:  col0→customer_code, col1→item_code, col2→mfg_number,
    # col3→model_number.
    raw = customer.get("dn_headers") or ""
    headers = raw.split(",")
    if len(headers) < 4:
        headers = [""] + headers  # old 3-field → prepend 客户号
    has_four = len([h for h in headers[:4] if h.strip()]) >= 4

    for ri, item in enumerate(items):
        row_num = 7 + ri
        if row_num > 18:
            break
        if has_four:
            _set(ws, "B{}".format(row_num), item.get("customer_code", ""))
            _set(ws, "C{}".format(row_num), item.get("item_code", ""))
            _set(ws, "D{}".format(row_num), item.get("mfg_number", ""))
            _set(ws, "E{}".format(row_num), item.get("model_number", ""))
            _set(ws, "F{}".format(row_num), item.get("product_name", ""))
            _set(ws, "G{}".format(row_num), item.get("color_name", ""))
            _set(ws, "H{}".format(row_num), item.get("quantity", 0))
            _set(ws, "I{}".format(row_num), item.get("unit_price", ""))
            amt = item.get("amount", 0)
            _set(ws, "J{}".format(row_num), amt if amt else "")
            _set(ws, "K{}".format(row_num), item.get("notes", ""))
        else:
            _set(ws, "B{}".format(row_num), item.get("item_code", ""))
            _set(ws, "C{}".format(row_num), item.get("model_number", ""))
            _set(ws, "D{}".format(row_num), item.get("product_name", ""))
            _set(ws, "E{}".format(row_num), item.get("color_name", ""))
            _set(ws, "F{}".format(row_num), item.get("quantity", 0))
            _set(ws, "G{}".format(row_num), item.get("unit_price", ""))
            amt = item.get("amount", 0)
            _set(ws, "H{}".format(row_num), amt if amt else "")
            _set(ws, "I{}".format(row_num), item.get("notes", ""))

    # 【核心修改】：删除了这里强行将页面缩放比例写死为 fitToWidth=1 的代码
    # 以保证程序完全尊重每个客户独立模板的原始打印设置。

    # Summary row
    total_amt = sum(
        (it.get("amount") or 0) for it in items
        if isinstance(it.get("amount"), (int, float)))
    from logic.formula import parse_formula
    total_pieces = 0
    for it in items:
        formula = it.get("quantity_formula") or ""
        if formula:
            s = formula.strip()
            if s.startswith("="):
                s = s[1:]
            r = parse_formula(s)
            if not r["error"] and r["piece_count"] > 0:
                total_pieces += r["piece_count"]
            else:
                total_pieces += 1
        else:
            total_pieces += 1
    if has_four:
        _set(ws, "I19", u"合计金额:")
        _set(ws, "J19", total_amt if total_amt else "")
        _set(ws, "K19", u"共{}个".format(total_pieces))
    else:
        _set(ws, "G19", u"合计金额:")
        _set(ws, "H19", total_amt if total_amt else "")
        _set(ws, "I19", u"共{}个".format(total_pieces))


def _format_header(left, right, right_pos):
    """Build a header string with *right* positioned at *right_pos*.

    Example:  ``_format_header('邮箱：...@126.COM', '客户订单号：HR001', 44)``
    """
    if len(left) >= right_pos - 2:
        left = left[:right_pos - 2]
    spaces = right_pos - len(left)
    if spaces < 1:
        spaces = 1
    return left + " " * spaces + right


def _clean_str(val):
    """Clean a string value: strip newlines / extra spaces / quotes."""
    if val is None:
        return None
    s = str(val)
    s = s.replace("\n", "").replace("\r", "").replace("\t", " ")
    s = s.strip().strip('"').strip("'").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    return s if s else None


def _set(ws, ref, val, _clean=True):
    """Set a cell value.

    Numbers (int / float) are written as-is so that Excel's number formatting
    is preserved.  Strings are cleaned first (unless ``_clean=False`` is
    passed for pre-formatted values like header cells).  None / empty strings
    are skipped.
    """
    if isinstance(val, (int, float)):
        ws[ref].value = val
        return
    if not _clean:
        # Used for header cells whose intentional multi-spacing must survive
        if val:
            ws[ref].value = val
        return
    cleaned = _clean_str(val)
    if cleaned is not None and cleaned != "":
        ws[ref].value = cleaned
