# -*- coding: utf-8 -*-
"""Convert existing .xls files to .xlsx format for system use."""
import os
import sys

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Read .xls with xlrd, write .xlsx with openpyxl."""
    import xlrd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    xls = xlrd.open_workbook(xls_path)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in xls.sheets():
        ws = wb.create_sheet(title=sheet.name)
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                val = sheet.cell(r, c).value
                ws.cell(row=r + 1, column=c + 1, value=val)
        # Basic formatting
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 36
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 14

    # Ensure output directory exists
    out_dir = os.path.dirname(xlsx_path)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir)

    wb.save(xlsx_path)
    return True


if __name__ == "__main__":
    # Convert all .xls files from 图例 to 送货单
    src_dir = os.path.join(os.path.dirname(_APP_DIR), u"图例")
    dst_dir = os.path.join(os.path.dirname(_APP_DIR), u"送货单")

    if not os.path.exists(dst_dir):
        os.makedirs(dst_dir)

    for fn in ["宏润", "雅兰"]:
        src = os.path.join(src_dir, "{}.xls".format(fn))
        dst = os.path.join(dst_dir, "{}.xlsx".format(fn))
        if os.path.exists(src):
            convert_xls_to_xlsx(src, dst)
            print(u"已转换: {} → {}".format(src, dst))
        else:
            print(u"未找到: {}".format(src))

    print(u"\n转换完成。")
